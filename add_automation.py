"""
Agrega la automatización pedida a Base_articulos.xlsx:
- F (Ingresos) y G (Egresos) en 'Base Articulos' pasan a ser fórmulas que
  buscan el Código en 'Ingresos y Egresos semanal' y traen el movimiento de
  esa semana. H ('=E+F-G', ya existente en el archivo) queda intacta y por
  lo tanto se recalcula sola.
- No se toca ninguna otra fórmula ni columna existente (regla de la skill de
  xlsx: al editar un archivo existente, respetar sus convenciones).
"""
from openpyxl import load_workbook
from openpyxl.comments import Comment

SRC = "Base_articulos.xlsx"
MOV_SHEET = "Ingresos y Egresos semanal"
MOV_LAST_ROW = 669  # Datos!A2:A669 en la hoja de movimientos (fila 1 = headers)

wb = load_workbook(SRC)
ba = wb["Base Articulos"]
assert ba.cell(row=1, column=8).value.strip() == "StkActual"
last_row = ba.max_row  # 3622

mov_codigo = f"'{MOV_SHEET}'!$A$2:$A${MOV_LAST_ROW}"
mov_ingresos = f"'{MOV_SHEET}'!$D$2:$D${MOV_LAST_ROW}"
mov_egresos = f"'{MOV_SHEET}'!$E$2:$E${MOV_LAST_ROW}"

for r in range(2, last_row + 1):
    ba.cell(row=r, column=6, value=f"=SUMIFS({mov_ingresos},{mov_codigo},B{r})")   # F: Ingresos
    ba.cell(row=r, column=7, value=f"=SUMIFS({mov_egresos},{mov_codigo},B{r})")    # G: Egresos

# Nota en los encabezados de F y G para que quede claro que no son de carga manual.
note = (
    "Esta columna se completa sola: busca el Código en la hoja "
    "'Ingresos y Egresos semanal' y trae el movimiento de esa semana. "
    "No escribas acá a mano — para actualizar el stock, cargá la semana en "
    "esa otra hoja."
)
ba["F1"].comment = Comment(note, "Amparo dashboard")
ba["G1"].comment = Comment(note, "Amparo dashboard")

wb.save(SRC)
print("Formulas agregadas en F2:F%d y G2:G%d" % (last_row, last_row))
