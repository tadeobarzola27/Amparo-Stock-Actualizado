"""
Agrega un chequeo de calidad de datos a Base_articulos.xlsx.

Motivo: al automatizar F/G de 'Base Articulos' con SUMIFS (add_automation.py)
se detectó que algunos códigos que registran movimientos en 'Ingresos y
Egresos semanal' NO existen en 'Base Articulos' -> esos Ingresos/Egresos no
se le descuentan/suman a ningún producto y el stock queda mal, en silencio.
Esto no es un caso hipotético: en el archivo real de Toni son 5 códigos y
~4589 unidades de Egresos "perdidos" en la carga actual.

Qué agrega este script (nada de lo existente se toca):
1. En 'Ingresos y Egresos semanal': columnas F y G (estaban vacías), un flag
   por fila + un contador auxiliar, ambos con fórmulas simples (COUNTIF /
   COUNTIFS), sin arrays.
2. Una hoja nueva al final, 'Códigos a revisar', con el resumen (cuántos
   códigos y cuánto Ingreso/Egreso quedan sin capturar) y el detalle de esos
   códigos, armado con INDEX/MATCH sobre el contador auxiliar (también sin
   arrays -> nada que dependa de spill/CSE).

Todas las funciones usadas están en la lista segura de la skill de xlsx:
COUNTIF, COUNTIFS, SUMPRODUCT, N, TRIM, INDEX, MATCH, IFERROR.
"""
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = "Base_articulos.xlsx"
MOV_SHEET = "Ingresos y Egresos semanal"
BASE_SHEET = "Base Articulos"
MOV_LAST_ROW = 669
BASE_LAST_ROW = 3622
MAX_ORPHANS_LISTADOS = 30  # hoy son 5; deja margen amplio para el futuro

wb = load_workbook(SRC)
mv = wb[MOV_SHEET]
ba = wb[BASE_SHEET]

assert mv.cell(row=1, column=1).value.strip() == "Codigo"
assert mv.cell(row=1, column=6).value is None, "F ya tiene contenido, no debería tocarse"
assert mv.cell(row=1, column=7).value is None, "G ya tiene contenido, no debería tocarse"

base_codigo_rng = f"'{BASE_SHEET}'!$B$2:$B${BASE_LAST_ROW}"
mov_codigo_rng = f"'{MOV_SHEET}'!$A$2:$A${MOV_LAST_ROW}"
mov_ingresos_rng = f"'{MOV_SHEET}'!$D$2:$D${MOV_LAST_ROW}"
mov_egresos_rng = f"'{MOV_SHEET}'!$E$2:$E${MOV_LAST_ROW}"
mov_desc_rng = f"'{MOV_SHEET}'!$C$2:$C${MOV_LAST_ROW}"

# ---------------------------------------------------------------------------
# 1) Columnas auxiliares F y G en la hoja de movimientos
# ---------------------------------------------------------------------------
mv["F1"] = "¿En Base Articulos?"
mv["G1"] = "Orden entre faltantes"
for col in ("F1", "G1"):
    mv[col].font = Font(bold=True)

for r in range(2, MOV_LAST_ROW + 1):
    mv.cell(
        row=r, column=6,
        value=f'=IF(TRIM(A{r})="","",IF(COUNTIF({base_codigo_rng},A{r})>0,"OK","Falta en Base Articulos"))',
    )
    mv.cell(
        row=r, column=7,
        value=f'=IF(F{r}="Falta en Base Articulos",COUNTIF($F$2:F{r},"Falta en Base Articulos"),"")',
    )

mv["F1"].comment = Comment(
    "Se completa sola. Indica si el Código de esta fila existe en la hoja "
    "'Base Articulos'. Si dice 'Falta en Base Articulos', el Ingreso/Egreso "
    "de esa fila no se está sumando a ningún stock -> ver la hoja "
    "'Códigos a revisar'.",
    "Amparo dashboard",
)
mv["G1"].comment = Comment(
    "Uso interno (para armar la hoja 'Códigos a revisar'). No hace falta "
    "leer esta columna a mano.",
    "Amparo dashboard",
)

# ancho razonable para que se lea el texto del flag
mv.column_dimensions["F"].width = 22
mv.column_dimensions["G"].width = 12

# ---------------------------------------------------------------------------
# 2) Hoja nueva 'Códigos a revisar'
# ---------------------------------------------------------------------------
SHEET_NAME = "Códigos a revisar"
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
alert = wb.create_sheet(SHEET_NAME)

TITLE_FILL = PatternFill("solid", fgColor="FFF4CC")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)

alert.merge_cells("A1:D1")
alert["A1"] = "⚠ Códigos con movimientos que no están en Base Articulos"
alert["A1"].font = Font(bold=True, size=13)
alert["A1"].fill = TITLE_FILL

alert.merge_cells("A2:D3")
alert["A2"] = (
    "Estos códigos tienen Ingresos y/o Egresos cargados en 'Ingresos y Egresos "
    "semanal' pero no existen en 'Base Articulos', así que ese movimiento no se "
    "descontó ni sumó a ningún stock. Para que dejen de aparecer acá, agregalos "
    "como fila nueva en 'Base Articulos' (con su Rubro, Descripción, etc.) — la "
    "próxima vez que se abra el archivo se van a sumar solos."
)
alert["A2"].alignment = Alignment(wrap_text=True, vertical="top")

labels = [
    ("A5", "Códigos sin catálogo:",
     f'=SUMPRODUCT((TRIM({mov_codigo_rng})<>"")*(COUNTIF({base_codigo_rng},{mov_codigo_rng})=0))'),
    ("A6", "Ingresos sin capturar:",
     f'=SUMPRODUCT((TRIM({mov_codigo_rng})<>"")*(COUNTIF({base_codigo_rng},{mov_codigo_rng})=0)*N({mov_ingresos_rng}))'),
    ("A7", "Egresos sin capturar:",
     f'=SUMPRODUCT((TRIM({mov_codigo_rng})<>"")*(COUNTIF({base_codigo_rng},{mov_codigo_rng})=0)*N({mov_egresos_rng}))'),
]
for cell, label, formula in labels:
    alert[cell] = label
    alert[cell].font = Font(bold=True)
    value_cell = f"B{cell[1:]}"
    alert[value_cell] = formula

alert.merge_cells("A9:D9")
alert["A9"] = "Detalle de códigos a agregar en Base Articulos"
alert["A9"].font = Font(bold=True)

headers = ["Codigo", "Descripcion", "Ingresos", "Egresos"]
for i, h in enumerate(headers):
    c = alert.cell(row=10, column=i + 1, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL

first_data_row = 11
for i in range(MAX_ORPHANS_LISTADOS):
    r = first_data_row + i
    rank = i + 1  # posición entre los faltantes (columna G de movimientos)
    match_expr = f"MATCH({rank},'{MOV_SHEET}'!$G$2:$G${MOV_LAST_ROW},0)"
    alert.cell(row=r, column=1, value=f'=IFERROR(INDEX({mov_codigo_rng},{match_expr}),"")')
    alert.cell(row=r, column=2, value=f'=IFERROR(INDEX({mov_desc_rng},{match_expr}),"")')
    alert.cell(row=r, column=3, value=f'=IFERROR(INDEX({mov_ingresos_rng},{match_expr}),"")')
    alert.cell(row=r, column=4, value=f'=IFERROR(INDEX({mov_egresos_rng},{match_expr}),"")')

alert.cell(row=first_data_row + MAX_ORPHANS_LISTADOS + 1, column=1,
           value=f"(Se muestran hasta {MAX_ORPHANS_LISTADOS} códigos. Si hubiera más, "
                 "el conteo de 'Códigos sin catálogo' de arriba lo va a reflejar igual.)")
alert.cell(row=first_data_row + MAX_ORPHANS_LISTADOS + 1, column=1).font = Font(italic=True, size=9, color="898781")

alert.column_dimensions["A"].width = 14
alert.column_dimensions["B"].width = 42
alert.column_dimensions["C"].width = 12
alert.column_dimensions["D"].width = 12

wb.save(SRC)
print("OK: columnas F/G agregadas en movimientos + hoja 'Códigos a revisar' creada.")
